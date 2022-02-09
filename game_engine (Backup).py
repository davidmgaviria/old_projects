# -*- coding: utf-8 -*-
"""
Created on Wed Nov 28 21:49:42 2018

@author: davidgaviria
"""

#%%
import openpyxl
import os
import xlwt
import math

workbook = xlwt.Workbook()
options = ('\nCommands:\n'
          'help - lists all possible commands \n'
          'update - takes you to the update screen \n'
          'battle - takes you to battle screen \n')
running = True
path = ('C:/Users/davidgaviria/Documents/My_Documents/Undecided_Game/Excel_Files')


def Run():
    global running
    
    while running:
        choice = input("Enter command: ")
        if choice=='help':
            print(options)
        elif choice=='update':
            filename = input('Filename (ALLP for all players, ALLA for all AI): ')
            if filename=='ALLP':
                p_files = ['zane.xlsx','teddy.xlsx','karan.xlsx','awad.xlsx','kobe.xlsx','ryan.xlsx'] 
                for name in p_files:
                    update(name)
            elif filename=='ALLA':
                ai_files = ['AI-kongan.xlsx', 'AI-aussie.xlsx', 'AI-guarani.xlsx','AI-siberian.xlsx','AI-suriname.xlsx','AI-swedish.xlsx'] 
                for name in ai_files:
                    update(name)
                    
            else:
                update(filename)
        elif choice=='battle':
           battle()   
        elif choice =='exit':
            print('Ending program')
            running = False
        else:
            print('Unknown command')
            
def getValue(spreadsheet, compare_value, min_r, min_c, max_c, column_letter_of_desired_variable ):
    for row in spreadsheet.iter_rows(min_row=min_r, min_col = min_c, max_col=max_c,):
        for cell in row:
            if cell.value is not None:
                if cell.value == compare_value:
                    desired_value = spreadsheet[column_letter_of_desired_variable+str(cell.row)].value
                    return desired_value
                    
                else:
                    pass
            else:
                print("Value not found")
                break    

############################################################################################################    
def update(filename):
    global chour, cmin
    
    xfile = openpyxl.load_workbook(os.path.join(path,filename))
    
    ##load all spreadsheets
    Scities = xfile['Cities']
    Sarmies = xfile['Armies']
    Sproduction = xfile['Production']
    Sconstruction = xfile['Construction']
    Smovement = xfile['Movement']
    Sinventory = xfile['Inventory']
    Sresearch = xfile['Research']
    Straining = xfile['Training']
    Sstats = xfile['Stats']
    
    ##Get time
    uhour = Sinventory['A2'].value
    umin = Sinventory['B2'].value
    
    Sinventory['A2'] = chour
    Sinventory['B2'] = cmin
    
    if(chour<uhour):
        etime = (1440-((uhour*60)+umin))+((chour*60)+cmin)
    else:
        etime = ((chour*60)+cmin)-((uhour*60)+umin)
    game_days = etime/28
    #print('Game Days: {}' .format(game_days))
    
    
    
    ##Get morale_modifier
    total = 0
    count = 0
    for row in Scities.iter_rows(min_row=2, min_col = 4, max_col=4):
        for cell in row:
            if cell.value is not None:
                total = total + cell.value  
                count = count + 1
                morale_modifier = (total/count)+0.5
            else:
                break
    
    
    
    #####################
    #####Cities Sheet####
    #####################
    
    ##Update population
    total_pop = 0
    for row in Scities.iter_rows(min_row=2, min_col = 3, max_col=3):
        for cell in row:
            if Scities['C'+str(cell.row)].value is not None:
                Scities['C'+str(cell.row)].value = int(Scities['C'+str(cell.row)].value*(((Sinventory['E18'].value/365)*game_days)+1))
                total_pop = total_pop+int(Scities['C'+str(cell.row)].value)
            else:
                break
    total_pop = total_pop + Sinventory['E14'].value + Sinventory['E15'].value 
            
    
                
    
         
    
    #####################
    #####Armies Sheet####
    #####################
    
    ##Update organization & combat effectiveness
    for row in Sarmies.iter_rows(min_row=2, min_col = 4, max_col=4):
        for cell in row:
            if cell.value is not None:
                count = 0
                level = Sarmies['G'+str(cell.row)].value
                while count<game_days and cell.value<1:
                    cell.value = cell.value+(level*0.1)
                if cell.value>1:
                    cell.value = 1
                
                Sarmies['F'+str(cell.row)].value = cell.value*Sarmies['C'+str(cell.row)].value
                
                
                
    
    
    ########################
    ####Prodcution Sheet####
    ########################
    
    #update production
    for row in Sproduction.iter_rows(min_row=2, min_col = 1, max_col=1):
        for cell in row:
            if cell.value is not None: 
                item_name = Sproduction['A'+str(cell.row)].value
                fac_assign = Sproduction['C'+str(cell.row)].value
                amount_per_day = 1/getValue(Sresearch, item_name, 3, 7, 7, 'H')
                oil_needed = getValue(Sresearch, item_name, 3, 7, 7, 'I')
                rubber_needed = getValue(Sresearch, item_name, 3, 7, 7, 'J')
                steel_needed = getValue(Sresearch, item_name, 3, 7, 7, 'K')
                aluminium_needed = getValue(Sresearch, item_name, 3, 7, 7, 'L')
                tungsten_needed = getValue(Sresearch, item_name, 3, 7, 7, 'M')
                chromium_needed = getValue(Sresearch, item_name, 3, 7, 7, 'N')
                Sproduction['B'+str(cell.row)].value = amount_per_day*fac_assign
                Sproduction['D'+str(cell.row)].value = oil_needed*fac_assign
                Sproduction['E'+str(cell.row)].value = rubber_needed*fac_assign
                Sproduction['F'+str(cell.row)].value = steel_needed*fac_assign
                Sproduction['G'+str(cell.row)].value = aluminium_needed*fac_assign
                Sproduction['H'+str(cell.row)].value = tungsten_needed*fac_assign
                Sproduction['I'+str(cell.row)].value = chromium_needed*fac_assign

                produced = Sproduction['B'+str(cell.row)].value*game_days*morale_modifier
                for row in Sinventory.iter_rows(min_row=24, min_col = 1, max_col=1):
                    for cell in row:
                        if cell.value==item_name:
                            Sinventory['B'+str(cell.row)].value = Sinventory['B'+str(cell.row)].value+int(produced)
                            break
                        else:
                            pass
                        
                            
                
                
                
            
    
    
    ########################
    ###Construction Sheet###
    ########################
    
    #update construction
    c_complete = 0
    for row in Sconstruction.iter_rows(min_row=2, min_col = 4, max_col=4):
        for cell in row:
            if cell.value is not None: 
                if cell.value == 'Completed':
                    c_complete +=1
                    continue
                else:
                    type_building = Sconstruction['A'+str(cell.row)].value
                    percent_left = 1-Sconstruction['D'+str(cell.row)].value
                    build_time = getValue(Sconstruction, type_building, 3, 10, 10, 'K')
                    if build_time is None:
                        print('Error with construction in: {}' .format(filename))
                    else:
                        left_to_build = percent_left*build_time
                        built_in_etime = Sconstruction['C'+str(cell.row)].value*game_days*morale_modifier
                        Sconstruction['D'+str(cell.row)].value = (left_to_build-built_in_etime)/build_time
            
                        if Sconstruction['D'+str(cell.row)].value<=0:
                            Sconstruction['D'+str(cell.row)].value = 'Completed'
                            c_complete +=1
            else:
                break
        
    
    
    
    
    
    #########################
    ###Movement Sheet###
    #########################
    
    ##Update location & time
    speed = 0
    m_complete = 0
    for row in Smovement.iter_rows(min_row=2, min_col = 5, max_col=5):
        for cell in row:
            if cell.value is not None:
                if cell.value == 'Arrived':
                    m_complete +=1
                    continue
                elif cell.value == '?':
                    start_tile = str(Smovement['C'+str(cell.row)].value)
                    end_tile = str(Smovement['D'+str(cell.row)].value)
                    speed = getValue(Sstats, Smovement['B'+str(cell.row)].value, 2, 1, 1, 'F' )
                    split = start_tile.find(',')
                    x1 = float(start_tile[:split])
                    y1 = float(start_tile[split+1:])
                    split = end_tile.find(',')
                    x2 = float(end_tile[:split])
                    y2 = float(end_tile[split+1:])
                    Smovement['F'+str(cell.row)].value = (math.sqrt(((x2-x1)**2)+((y2-y1)**2)))*121
                    Smovement['E'+str(cell.row)].value = 28*(Smovement['F'+str(cell.row)].value/(speed*12))
                else:
                    speed = getValue(Sstats, Smovement['B'+str(cell.row)].value, 2, 1, 1, 'F' )
                    
                Smovement['F'+str(cell.row)].value = Smovement['F'+str(cell.row)].value-(game_days*speed*12)
                Smovement['E'+str(cell.row)].value = 28*(Smovement['F'+str(cell.row)].value/(speed*12))
                if Smovement['E'+str(cell.row)].value<=0:
                    Smovement['E'+str(cell.row)].value = 'Arrived'
                    Smovement['C'+str(cell.row)].value = Smovement['D'+str(cell.row)].value
                    m_complete +=1
         

                        



   
    
    
    ####################
    ###Research Sheet###
    ####################
    
    ##Update research
    r_complete = 0
    for row in Sresearch.iter_rows(min_row=2, min_col = 3, max_col=3):
        for cell in row:
            if cell.value is not None:
                if cell.value == 'Completed':
                    r_complete +=1
                Sresearch['C'+str(cell.row)].value = Sresearch['C'+str(cell.row)].value-game_days
                if Sresearch['C'+str(cell.row)].value<=0:
                    Sresearch['C'+str(cell.row)].value = 'Completed'
                    r_complete +=1
                else:
                    pass
            else:
                break
            
            
        
        
        
        
        
    ####################
    ###Training Sheet###
    ####################
    
    ##Update training
    t_complete = 0
    for row in Straining.iter_rows(min_row=2, min_col = 5, max_col=5):
        for cell in row:
            if cell.value is not None:
                if cell.value == 'Trained':
                     t_complete +=1
                     continue
                elif cell.value == '?':
                    name = str(Straining['B'+str(cell.row)].value)
                    level = Straining['D'+str(cell.row)].value
                    time = getValue(Sstats, name, 2, 1, 1, 'I' )
                    Straining['E'+str(cell.row)].value = 28*level*time
                else: 
                    pass
                    
                Straining['E'+str(cell.row)].value =  float(Straining['E'+str(cell.row)].value) - (28*game_days)
                if Straining['E'+str(cell.row)].value<=0:
                    Straining['E'+str(cell.row)].value = 'Trained'
                    t_complete +=1

                        
                    
        
        
        
        
                  
    #####################
    ###Inventory Sheet###
    #####################
    
    #Get population
    Sinventory['E11'].value = total_pop
    Sinventory['E12'].value = total_pop*Sinventory['E19'].value
    
    ##Get average morale
    total = 0
    count = 0
    for row in Scities.iter_rows(min_row=2, min_col = 4, max_col=4):
        for cell in row:
            if cell.value is not None:
                total = total + cell.value  
                count = count + 1
                Sinventory['E17'] = (total/count)
            else:
                break
            

    #civilian & military factory count
    avail_civ = 0
    damag_civ = 0
    avail_mil = 0
    damag_mil = 0
    for row in Scities.iter_rows(min_row=2, min_col = 5, max_col=5):
        for cell in row:
            if cell.value is not None:
                avail_civ = avail_civ + Scities['E'+str(cell.row)].value

    for row in Scities.iter_rows(min_row=2, min_col = 7, max_col=7):
        for cell in row:
            if cell.value is not None:
                damag_civ = damag_civ + Scities['G'+str(cell.row)].value
                
    for row in Scities.iter_rows(min_row=2, min_col = 6, max_col=6):
        for cell in row:
            if cell.value is not None:
                avail_mil = avail_mil + Scities['F'+str(cell.row)].value
    
    for row in Scities.iter_rows(min_row=2, min_col = 8, max_col=8):
        for cell in row:
            if cell.value is not None:
                damag_mil = damag_mil + Scities['H'+str(cell.row)].value
    
    Sinventory['H11'].value = avail_civ + damag_civ
    Sinventory['H12'].value = avail_civ
    Sinventory['H13'].value = avail_mil + damag_mil
    Sinventory['H14'].value = avail_mil
    
    
    xfile.save(os.path.join(path,filename))
    print('\n#################################\n'
          'File: {}'.format(filename))
    if c_complete > 0:
        print('{} construction(s) completed' .format(c_complete))
    if m_complete > 0:
        print('{} unit(s) have arrived at their destinations' .format(m_complete))
    if r_complete > 0:
        print('{} item(s) have been researched' .format(r_complete))
    if t_complete > 0:
        print('{} unit(s) have finished training' .format(t_complete))
    print('Successfully update file\n'
          '#################################\n')
          
          
          
#########################################################################################
def battle():
    
    battlefile = openpyxl.load_workbook(os.path.join(path,'zzz_battle.xlsx'))
    Sland = battlefile['Land']
    Snaval = battlefile['Naval']
    Sair = battlefile['Air']
    
    player1 = input("Player 1's File: ")
    update(player1)
    player2 = input("Player 2's File: ")
    update(player2)
    print('Returning to battle...')
    
    p1file = openpyxl.load_workbook(os.path.join(path,player1))
    p2file = openpyxl.load_workbook(os.path.join(path,player2))
    
    ##load all spreadsheets
    P1cities = p1file['Cities']
    P1armies = p1file['Armies']
    P1inventory = p1file['Inventory']
    P1stats = p1file['Stats']
    
    P2cities = p2file['Cities']
    P2armies = p2file['Armies']
    P2inventory = p2file['Inventory']
    P2stats = p2file['Stats']
    
    print('Successfully completed')
    
    
    
    
    
    





    
    
    
    
###########################################################################################   
global chour, cmin
ctime = input('Current Time: ')
spot = ctime.find(':')
chour = int(ctime[:spot])
cmin = int(ctime[spot+1:])
print(options)
Run()
#%%%
