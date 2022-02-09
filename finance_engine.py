# -*- coding: utf-8 -*-
"""
Created on Tue Jan  8 08:41:39 2019

@author: davidgaviria
"""
#%%
import openpyxl
import os
import xlwt
import math

workbook = xlwt.Workbook()
path = ('C:/Users/David Gaviria/Documents/')
filename = 'Finances.xlsx'


xfile = openpyxl.load_workbook(os.path.join(path,filename))
  
Overview = xfile['Overview']  
FP = xfile['Food & Personal Care']
School = xfile['School Supplies']
Auto = xfile['Auto & Transportation']
Activities = xfile['Activities']
Gifts = xfile['Gifts']
Other = xfile['Other']
WDC = xfile['WDC']
Loans = xfile['Loans']


#get totals from all categories

Overview['F12'].value = 0
for row in FP.iter_rows(min_row=4, min_col=3, max_col=3):
        for cell in row:
            if cell.value is not None:
                Overview['F12'].value += cell.value
                
                
Overview['F13'].value = 0
for row in School.iter_rows(min_row=4, min_col=3, max_col=3):
        for cell in row:
            if cell.value is not None:
                Overview['F13'].value += cell.value
        

Overview['F14'].value = 0
for row in Auto.iter_rows(min_row=4, min_col=3, max_col=3):
        for cell in row:
            if cell.value is not None:
                Overview['F14'].value += cell.value
                

Overview['F15'].value = 0               
for row in Activities.iter_rows(min_row=4, min_col=3, max_col=3):
        for cell in row:
            if cell.value is not None:
                Overview['F15'].value += cell.value
               
                
Overview['F16'].value = 0               
for row in Gifts.iter_rows(min_row=4, min_col=3, max_col=3):
        for cell in row:
            if cell.value is not None:
                Overview['F16'].value += cell.value
                
                
Overview['F17'].value = 0               
for row in Other.iter_rows(min_row=4, min_col=3, max_col=3):
        for cell in row:
            if cell.value is not None:
                Overview['F17'].value += cell.value
                

withdrawal = 0
deposit = 0
cash = 0
for row in WDC.iter_rows(min_row=4, min_col=4, max_col=4):
        for cell in row:
            if cell.value is not None:
                if WDC['B'+str(cell.row)].value == 'W':
                    withdrawal = withdrawal+cell.value
                elif WDC['B'+str(cell.row)].value == 'D':
                    deposit = deposit+cell.value
                elif WDC['B'+str(cell.row)].value == 'C':
                    cash = cash+cell.value
                else:
                    print('Unknown type in WDC, cell B'+str(cell.row))
                    
                    
loans = 0               
Overview['I4'].value = 0        #clears money owned
for row in Loans.iter_rows(min_row=4, min_col=3, max_col=3):
        for cell in row:
            if cell.value is not None:
               loans += cell.value
               if Loans['D'+str(cell.row)].value == 'No':
                   Overview['I4'].value += cell.value    
               
               
                
#modify balance
Overview['B14'].value = Overview['F12'].value+Overview['F13'].value+Overview['F14'].value+Overview['F15'].value+Overview['F16'].value+Overview['F17'].value+loans   #get money spent
Overview['B12'].value = Overview['B11'].value-Overview['B14'].value-withdrawal+deposit    #get current amount of money
Overview['B15'].value = Overview['B12'].value-Overview['B11'].value    #net change

Overview['F4'].value = math.ceil(cash)
Overview['F5'].value = Overview['F4'].value+Overview['B14'].value

Overview['F18'].value = Overview['B14'].value-loans
for row in Overview.iter_rows(min_row=12,max_row=17,min_col=7,max_col=7):
    for cell in row:
        if Overview['F'+str(cell.row)].value!=0:
            cell.value = (Overview['F'+str(cell.row)].value/Overview['F18'].value)
        else:
            pass


xfile.save(os.path.join(path,filename))        
print('Done')